import pandas
import xlrd
import xlwt
from openpyxl import Workbook
from tempfile import TemporaryFile

col_num_vuids = 5

print("Loading workbooks")
previous_voters = xlrd.open_workbook('Pct 3,5,10 (list)2019.xlsx',{'strings_to_numbers':True})
voters = xlrd.open_workbook('Voting History clean.xlsx')

book = xlwt.Workbook()
nonvoter_sheet = book.add_sheet('nonvoters')

print("Workbooks loaded")
sname = previous_voters.sheet_names()
vsname = voters.sheet_names()

previous_voters_sheet = previous_voters.sheet_by_name(sname[0])
voters_sheet = voters.sheet_by_name(vsname[0])

vuids = []
for i in range(voters_sheet.nrows):
    vuids.append(voters_sheet.cell(i,col_num_vuids).value)

print("prev Voters: " , previous_voters_sheet.nrows)
print("recent Voters: ", voters_sheet.nrows)

print("finding nonvoters...")
nonvoters = []
col = 0
n_nonvoters = 0

#Add column labels to worksheet
for col in range(previous_voters_sheet.ncols):
    cell_data = previous_voters_sheet.cell_value(0,col)
    if(cell_data):
        nonvoter_sheet.write(0,col,cell_data)
        cell_data=''

print("Nonvoter sheet:" ,nonvoter_sheet.row(0))

#Compare unique voter id, if recent voter not found add to nonvoters worksheet
row_index = 1
for i in range(previous_voters_sheet.nrows):
    found = False
    for j in range(voters_sheet.nrows):
        if(previous_voters_sheet.cell(i,col_num_vuids-1).value == vuids[j]):
            found = True

    if(not found):
        n_nonvoters+=1
        for col in range(previous_voters_sheet.ncols):
            cell_data = previous_voters_sheet.cell_value(i,col)
            if(cell_data):
                nonvoter_sheet.write(row_index,col,cell_data)
                cell_data=''
        row_index +=1


print("nonvoters found: ", n_nonvoters)
name = 'nonvoters2019.xls'
book.save(name)
book.save(TemporaryFile())
print("file saved as ", name)
